import os
import openpyxl
import io
import xlwt
import json
import datetime
from django import forms
import pandas as pd
from xlrd import open_workbook
from django.shortcuts import render
from django.http import HttpResponse
from .models import Calltrack_lite
from django.http import JsonResponse
from .models import Departs, Category_managers
import datetime


def index(request):
    return render(request, 'bot_back_end/index.html')


def about(request):
    return render(request, 'bot_back_end/about.html')


def export_calltrack_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="calltrack.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('calltrack')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['id', 'время', 'фраза клиента', 'вх.номер', 'регион',
               'доб.номер', 'ключевые слова']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = Calltrack_lite.objects.all().values_list(
        'id_rec', 'date_time_calling', 'text', 'innumber', 'region',
        'dial_route', 'key_words')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if isinstance(row[col_num], datetime.datetime):
                date_time = row[col_num].strftime('%Y-%m-%d %H:%M:%S')
                ws.write(row_num, col_num, date_time, font_style)
            elif col_num == 4:
                name_dial = get_region_name(row[4])
                ws.write(row_num, col_num, name_dial, font_style)
            else:
                ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response


def get_depart_or_name_manager(request):
    dial_route = request.GET.get('dial_route')

    if Departs.objects.filter(depart_added_phone_number=dial_route):
        name = Departs.objects.get(depart_added_phone_number=dial_route).depart_name
    elif Category_managers.objects.filter(phone_number=dial_route):
        name = Category_managers.objects.filter(phone_number=dial_route).values_list('name_man')[0][0]
    else:
        name = dial_route
    response = JsonResponse({'name': name})
    return response


def get_depart_or_name_manager_func(dial_route):

    if Departs.objects.filter(depart_added_phone_number=dial_route):
        name = Departs.objects.get(depart_added_phone_number=dial_route).depart_name
    elif Category_managers.objects.filter(phone_number=dial_route):
        name = Category_managers.objects.filter(phone_number=dial_route).values_list('name_man')[0][0]
    else:
        name = dial_route
    print('name:', name)
    return name


def get_region_name(city_code):
    name_filials = {
                101: 'Республика Адыгея',
                102: 'Республика Башкортостан',
                103: 'Республика Бурятия',
                104: 'Республика Алтай',
                105: 'Республика Дагестан',
                106: 'Республика Ингушетия',
                107: 'Кабардино-Балкарская Республика',
                108: 'Республика Калмыкия',
                109: 'Карачаево-Черкесская Республика',
                110: 'Республика Карелия',
                111: 'Республика Коми',
                112: 'Республика Марий Эл',
                113: 'Республика Мордовия',
                114: 'Республика Саха (Якутия)',
                115: 'Республика Северная Осетия - Алания',
                116: 'Республика Татарстан (Татарстан)',
                117: 'Республика Тыва',
                118: 'Удмуртская Республика',
                119: 'Республика Хакасия',
                120: 'Чеченская Республика',
                121: 'Чувашская Республика - Чувашия',
                122: 'Алтайский край',
                123: 'Краснодарский край',
                124: 'Красноярский край',
                125: 'Приморский край',
                126: 'Ставропольский край',
                127: 'Хабаровский край',
                128: 'Амурская область',
                129: 'Архангельская область',
                130: 'Астраханская область',
                131: 'Белгородская область',
                132: 'Брянская область',
                133: 'Владимирская область',
                134: 'Волгоградская область',
                135: 'Вологодская область',
                136: 'Воронежская область',
                137: 'Ивановская область',
                138: 'Иркутская область',
                139: 'Калининградская область',
                140: 'Калужская область',
                141: 'Камчатская область',
                142: 'Кемеровская область',
                143: 'Кировская область',
                144: 'Костромская область',
                145: 'Курганская область',
                146: 'Курская область',
                147: 'Ленинградская область',
                148: 'Липецкая область',
                149: 'Магаданская область',
                150: 'Московская область',
                151: 'Мурманская область',
                152: 'Нижегородская область',
                153: 'Новгородская область',
                154: 'Новосибирская область',
                155: 'Омская область',
                156: 'Оренбургская область',
                157: 'Орловская область',
                158: 'Пензенская область',
                159: 'Пермская область',
                160: 'Псковская область',
                161: 'Ростовская область',
                162: 'Рязанская область',
                163: 'Самарская область',
                164: 'Саратовская область',
                165: 'Сахалинская область',
                166: 'Свердловская область',
                167: 'Смоленская область',
                168: 'Тамбовская область',
                169: 'Тверская область',
                170: 'Томская область',
                171: 'Тульская область',
                172: 'Тюменская область',
                173: 'Ульяновская область',
                174: 'Челябинская область',
                175: 'Читинская область',
                176: 'Ярославская область',
                177: 'Москва',
                178: 'Санкт-Петербург',
                179: 'Еврейская автономная область',
                180: 'Агинский Бурятский автономный округ',
                181: 'Коми-Пермяцкий автономный округ',
                182: 'Корякский автономный округ',
                183: 'Ненецкий автономный округ',
                184: 'Таймырский (Долгано-Ненецкий) автономный округ',
                185: 'Усть-Ордынский Бурятский автономный округ',
                186: 'Ханты-Мансийский автономный округ',
                187: 'Чукотский автономный округ',
                188: 'Эвенкийский автономный округ',
                189: 'Ямало-Ненецкий автономный округ',
                191: 'Крым',
                198: 'Санкт-Петербург',
                199: 'Московская обл',
                800: 'Бесплатный номер',
                1:   'США и Канада',
                2:   'Африка',
                3:   'Европа',
                4:   'Европа',
                5:   'Южная, центральная Америка и Мексика',
                6:   'Австралия и Океания',
                7:   'СНГ',
                8:   'Китай, Япония, Корея и др. страны Азии',
                9:   'Индия, Турция, Армения, и др. страны Азии',
                863: 'Красноярский край',
                861: 'Красноярский край',
                383: 'Новосибирская область'
    }


    name_filial = name_filials[city_code]
    return name_filial


################################
#Формирование файла с ответом###
################################


def get_wb_xls(path, filename =  'calltrack.xlsx'):
    path_file = os.path.join(path, filename)
    wb = open_workbook(path_file)
    sheet_names = wb.sheet_names()
    xl_sheet = wb.sheet_by_name(sheet_names[0])
    rows = []
    for row_idx in range(0, xl_sheet.nrows):    # Iterate through rows
        rows.append([xl_sheet.cell(row_idx, 1), xl_sheet.cell(row_idx, 3), xl_sheet.cell(row_idx, 6)])
    return rows


def get_wb_csv(path, filename = 'mango.csv'):
    path_file = os.path.join(path, filename)
    df = pd.read_csv(path_file, encoding='cp1251', delimiter=';', header=1)
    df = df[['Кто звонил', 'Участники']]
    df = df[df['Участники'] != '-']
    rows = []


    for row in zip(df['Кто звонил'], df['Участники']):
        s = row[1].split(',')
        rows.append([row[0], s[-1].lstrip()])
    return rows


def get_z(x, y):
    z = list()
    for rowx in x:
        for rowy in y:
            telx = str(rowx[0])
            tely = rowy[1].value
            keys = rowy[2].value
            if telx == tely and keys != 'Error' and keys != '':
                z.append([rowy[0].value, rowx[1], tely, keys])
    return z


def export_missed_calls(z, path, filename="missedCalls.xls"):
    path_file = os.path.join(path, filename)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="missedCalls.xlsx"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('missedCalls')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Дата и время', 'номер', 'имя менежера', 'кл. слова']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = z

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(path_file)
    wb.save(response)
    return response


def save_file(f, itemfile):
    with open(itemfile, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def get_missed_calls(request):
    path = os.path.join('/home/asterisk/bot_django/bot_back_end/media')
    item_file_calltrack = os.path.join(path, 'calltrack.xlsx')
    item_file_mango = os.path.join(path, 'mango.csv')
    if request.method == 'POST':
        calltrack =request.FILES['calltrack']
        mango = request.FILES['mango']
        save_file(calltrack, item_file_calltrack)
        save_file(mango, item_file_mango)

        x = get_wb_csv(path)
        y = get_wb_xls(path)
        z = get_z(x, y)
        response = export_missed_calls(z, path)
        return response